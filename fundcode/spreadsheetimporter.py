import openpyxl
import datetime
import pyodbc
from fund import ORACLESTRING
from fundamountparser import oracledatebuilder

path = "I:\\Data\\Actuary\\Risk Management - Equity Market\\Fund Analysis\\Return Data\\PLFA\\forupload\\"

FILENAMES = [path+"Individual Fund Returns 1-2-02 to 12-31-13.xlsx",
             path+"PSF and RE Returns 2-26-92 to 12-31-13.xlsx"]

def importssdata(filenames):
    importedfunds = []
    cnxn = pyodbc.connect(ORACLESTRING)
    c = cnxn.cursor()
    c.execute('delete from funddatanew;')
    c.commit()
    for filename in filenames:
        wb = openpyxl.load_workbook(filename)
        ws = wb.get_sheet_by_name('Fund Returns - Daily')
        colnum = 2
        while not(ws.cell(row=1,column=colnum).value==None):
            dates, returns = [],[]
            rownum = 3
            totalreturn = 1.0
            while ws.cell(row=rownum,column=colnum).value==None:
                rownum+=1
            while not(ws.cell(row=rownum,column=colnum).value==None):
                divisor = 100.0 if filename == FILENAMES[0] else 1.0
                returnval = ws.cell(row=rownum,column=colnum).value/divisor
                totalreturn *= (1.0+returnval)
                if returnval!=0.0:
                    dte = ws.cell(row=rownum,column=1).value
                    dates.append(dte)
                    returns.append(totalreturn)           
                rownum+=1
            if not(ws.cell(row=0,column=colnum).value=='SKIP'):
                fundnum = int(ws.cell(row=0,column=colnum).value)
                if fundnum in importedfunds:
                    print 'DUPLICATE FUND NUMBER - ' + str(fundnum)
                else:
                    importedfunds.append(fundnum)
                print fundnum
                c.execute('delete from funddatanew where fundnum='+str(fundnum)+';')
                c.commit()
                for date, returnval in zip(dates,returns):
                    sql = "insert into funddatanew values(101,'BASE',"+oracledatebuilder(date)+","+str(fundnum)+","+str(returnval)+');'
                    c.execute(sql)
                c.commit()
            colnum+=1



if __name__=='__main__':
    importssdata(filenames=FILENAMES)